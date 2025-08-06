from openpyxl import load_workbook
from typing import Dict, List
from openpyxl.utils import range_boundaries

FIELD_DESC = {
    "name": "算法/组件等的英文名称",
    "primitive": "高级密码系统和协议中使用的密码构建块。原语代表不同的密码例程：确定性随机位生成器（drbg，例如 NIST SP800-90A-r1 中的 CTR_DRBG）、消息认证码（mac，例如 HMAC-SHA-256）、分组密码（例如 AES）、流密码（例如 Salsa20）、签名（例如 ECDSA）、哈希函数（例如 SHA-256）、公钥加密方案（pke，例如 RSA）、扩展输出函数（xof，例如 SHAKE256）、密钥派生函数（例如 pbkdf2）、密钥协商算法（例如 ECDH）、密钥封装机制（例如 ML-KEM）、认证加密（ae，例如 AES-GCM）以及多种算法的组合（组合器，例如 SP800-56Cr2）。",
    "parameter_set": "加密算法参数集的标识符。示例：在 AES128 中，“128”表示密钥长度（以位为单位）；在 SHA256 中，“256”表示摘要长度；在 SHAKE128 中，“128”表示其最高安全级别（以位为单位）；“SHA2-128s”表示 SLH-DSA (FIPS205) 中使用的参数集。",
    "mode": "使用加密算法（分组密码）的操作模式。",
    "platform": "算法实现的目标平台。该实现可以是“通用的”，可在任何平台上运行，也可以针对特定平台运行。",
    "execution_environment": "实现算法的目标和执行环境。",
    "padding": "用于加密算法的填充方案（可忽略）。",
    "functions": "密码算法所实现的密码功能。",
    "oid": "ASN.1 OID 或自定义标识符, 对象标识符（OID）。",
    "classical_security_level": "加密算法提供的经典安全级别（以位为单位）。",
    "nist_level(Not necessary)": "NIST 安全强度类别，定义见 https://csrc.nist.gov/projects/post-quantum-cryptography/post-quantum-cryptography-standardization/evaluation-criteria/security-(evaluation-criteria)。值为 0 表示不满足任何类别。（可忽略）",
    "bom-ref": "SBOM/CBOM 内部引用 ID, 可选标识符，可用于在 BOM 中的其他位置引用注释。每个 bom-ref 在 BOM 中必须是唯一的。"
    # ....
}


def describe_field(field: str) -> str:
    """根据字段名返回中文描述；未知字段返回原字符串。"""
    return FIELD_DESC.get(field.strip().lower(), field)


def extract_cbom_headers(file_path: str) -> Dict[str, List[str]]:
    """
    返回格式：
    {
        "CBOM-App": ["标题1", "标题2", ...],
        "CBOM-Data": ["标题A", "标题B", ...],
        ...
    }
    """
    wb = load_workbook(file_path, data_only=True)
    result = {}

    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith("CBOM"):
            continue

        ws = wb[sheet_name]

        merged_tops = set()
        for merged_range in ws.merged_cells.ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    if row == min_row and col == min_col:
                        continue
                    merged_tops.add((row, col))

        headers = []
        for row_num in [5]:
            for col in range(1, ws.max_column + 1):
                if (row_num, col) in merged_tops:
                    continue
                cell_value = ws.cell(row=row_num, column=col).value
                if cell_value is not None:
                    headers.append(str(cell_value).strip())

        result[sheet_name] = headers

    return result


# 示例用法
if __name__ == "__main__":
    file_path = "18031-CBOM.xlsx"
    headers_dict = extract_cbom_headers(file_path)
    for sheet, headers in headers_dict.items():
        print(f"[+]CBOM: {sheet}")
        print("[+]headers:", headers)
        print("-" * 80)
    h = "parameter_set"
    print(f"[+]字段描述")
    print(f"    {h:20} --> {describe_field(h)}")
